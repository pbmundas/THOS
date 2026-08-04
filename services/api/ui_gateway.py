"""Authenticated SOCmate UI gateway for THOS hunt and report workflows."""
from __future__ import annotations

import asyncio
import base64
from contextlib import asynccontextmanager
from datetime import datetime, timezone
import hashlib
import hmac
from html import escape
from io import BytesIO
import json
import logging
import os
from pathlib import Path
import re
import secrets
import sys
import time
from typing import AsyncIterator
import uuid

import httpx
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, JSONResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field, model_validator
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    KeepTogether,
    PageBreak,
    Paragraph,
    Preformatted,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from services.api import control_plane
from services.hunting.hypothesis_catalog import normalize_hypothesis
from services.runtime_config import get_value, read_config
from services.security.configuration import required_secret


class _JsonFormatter(logging.Formatter):
    def format(self, record: logging.LogRecord) -> str:
        payload = {
            "timestamp": datetime.fromtimestamp(record.created, timezone.utc).isoformat(),
            "level": record.levelname,
            "service": "thos-ui",
            "logger": record.name,
            "message": record.getMessage(),
        }
        if record.exc_info:
            payload["exception"] = self.formatException(record.exc_info)
        return json.dumps(payload, default=str)


_handler = logging.StreamHandler(sys.stdout)
_handler.setFormatter(_JsonFormatter())
logging.getLogger().handlers.clear()
logging.getLogger().addHandler(_handler)
logging.getLogger().setLevel(os.environ.get("LOG_LEVEL", "INFO").upper())
logging.getLogger("httpx").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)

ORCHESTRATOR_URL = os.environ.get("ORCHESTRATOR_URL", "http://orchestrator:8200").rstrip("/")
ORCHESTRATOR_API_KEY = required_secret("ORCHESTRATOR_API_KEY")
REPORTS_DIR = Path(os.environ.get("REPORTS_DIR", "/data/reports"))
FORENSIC_ROOT = Path(os.environ.get("FORENSIC_ROOT", "/data/log_sources/forensic"))
STATIC_DIR = Path(os.environ.get("UI_STATIC_DIR", "/app/static"))
_UPSTREAM_HEADERS = {"Authorization": f"Bearer {ORCHESTRATOR_API_KEY}"}
FORENSIC_MAX_FILE_BYTES = int(os.environ.get("FORENSIC_MAX_FILE_BYTES", str(20 * 1024 * 1024 * 1024)))
FORENSIC_MAX_CASE_BYTES = int(os.environ.get("FORENSIC_MAX_CASE_BYTES", str(100 * 1024 * 1024 * 1024)))
_forensic_intake_lock = asyncio.Lock()


def _parse_accounts() -> list[tuple[str, str]]:
    raw = os.environ.get("CHATUI_USERS", "").strip()
    accounts: list[tuple[str, str]] = []
    if raw:
        for entry in raw.split(","):
            username, separator, password = entry.strip().partition(":")
            if (
                separator
                and username
                and len(password) >= 12
                and "change_me" not in password.casefold()
            ):
                accounts.append((username, password))
            else:
                raise RuntimeError(
                    "CHATUI_USERS entries require a username and a unique "
                    "password of at least 12 characters"
                )
    else:
        username = os.environ.get("CHATUI_USERNAME", "").strip()
        password = os.environ.get("CHATUI_PASSWORD", "")
        if not username:
            raise RuntimeError(
                "CHATUI_USERNAME or CHATUI_USERS must be explicitly configured"
            )
        if len(password) < 12 or "change_me" in password.casefold():
            raise RuntimeError(
                "CHATUI_PASSWORD must be explicitly configured with at least "
                "12 characters"
            )
        accounts.append((username, password))
    if not accounts:
        raise RuntimeError("no valid UI accounts configured")
    return accounts


UI_ACCOUNTS = _parse_accounts()
control_plane.seed_users(UI_ACCOUNTS)
control_plane.seed_ioc_sources()
SESSION_COOKIE = "thos_session"
SESSION_TTL_SECONDS = max(900, int(os.environ.get("CHATUI_SESSION_TTL_SECONDS", "43200")))
SESSION_SECURE_COOKIE = os.environ.get("CHATUI_SECURE_COOKIE", "0").strip().lower() in {"1", "true", "yes"}
SESSION_SECRET = required_secret("CHATUI_SESSION_SECRET").encode("utf-8")
@asynccontextmanager
async def _lifespan(_app: FastAPI):
    control_plane.start_scheduler()
    try:
        yield
    finally:
        await control_plane.stop_scheduler()


app = FastAPI(
    title="THOS SOCmate UI",
    version="1.0.0",
    docs_url=None,
    redoc_url=None,
    lifespan=_lifespan,
)

_SPA_STATIC_ROUTES = {
    "/", "/overview", "/log-search", "/risks", "/detections", "/hunt-board",
    "/forensic", "/forensic/evidence", "/forensic/yara",
    "/threat-intelligence", "/reports", "/integrations", "/configuration",
    "/help", "/hypotheses/new", "/workspace",
}
_SPA_DYNAMIC_ROUTES = (
    re.compile(r"^/detections/[A-Za-z0-9._:-]{1,128}$"),
    re.compile(r"^/hunt-board/[0-9a-fA-F-]{36}$"),
    re.compile(r"^/forensic/(?:evidence|yara)/[0-9a-fA-F-]{36}$"),
    re.compile(r"^/reports/[A-Za-z0-9._-]{1,255}$"),
    re.compile(r"^/configuration/[a-z][a-z-]{0,31}$"),
)


def _is_spa_route(path: str) -> bool:
    normalized = path.rstrip("/") or "/"
    return normalized in _SPA_STATIC_ROUTES or any(
        pattern.fullmatch(normalized) for pattern in _SPA_DYNAMIC_ROUTES
    )


async def _record_audit_event(payload: dict) -> None:
    try:
        async with httpx.AsyncClient(timeout=httpx.Timeout(5, connect=2)) as client:
            await client.post(
                f"{ORCHESTRATOR_URL}/audit/events",
                headers=_UPSTREAM_HEADERS,
                json=payload,
            )
    except Exception:  # noqa: BLE001 - audit transport cannot break the UI
        logger.warning("could not persist UI audit event", exc_info=True)


def _valid_account(supplied_user: str, supplied_password: str) -> str:
    user = control_plane.authenticate(supplied_user, supplied_password)
    return str(user.get("username", "")) if user else ""


def _session_token(username: str) -> str:
    payload = json.dumps(
        {"sub": username, "exp": int(datetime.now(timezone.utc).timestamp()) + SESSION_TTL_SECONDS},
        separators=(",", ":"),
    ).encode("utf-8")
    encoded = base64.urlsafe_b64encode(payload).rstrip(b"=").decode("ascii")
    signature = hmac.new(SESSION_SECRET, encoded.encode("ascii"), hashlib.sha256).hexdigest()
    return f"{encoded}.{signature}"


def _session_user(token: str) -> str:
    encoded, separator, supplied_signature = token.partition(".")
    if not separator:
        return ""
    expected = hmac.new(SESSION_SECRET, encoded.encode("ascii"), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(supplied_signature, expected):
        return ""
    try:
        padded = encoded + "=" * (-len(encoded) % 4)
        payload = json.loads(base64.urlsafe_b64decode(padded).decode("utf-8"))
        username = str(payload.get("sub", ""))
        expires = int(payload.get("exp", 0))
    except (ValueError, TypeError, UnicodeDecodeError, json.JSONDecodeError):
        return ""
    if expires <= int(datetime.now(timezone.utc).timestamp()):
        return ""
    user = control_plane.get_user(username)
    return username if user and user.get("enabled", True) else ""


@app.middleware("http")
async def require_ui_auth(request: Request, call_next):
    public_path = (
        request.url.path in {"/", "/index.html", "/health", "/api/auth/login", "/api/auth/logout"}
        or request.url.path.startswith("/assets/")
        or _is_spa_route(request.url.path)
    )
    if public_path:
        return await call_next(request)
    valid_user = _session_user(request.cookies.get(SESSION_COOKIE, ""))
    if not valid_user:
        return JSONResponse(status_code=401, content={"detail": "authentication required"})
    user = control_plane.get_user(valid_user) or {}
    request.state.analyst = valid_user
    request.state.display_name = user.get("display_name") or valid_user
    request.state.role = user.get("role", "Expert")
    request.state.permissions = set(user.get("permissions", []))
    started_at = time.perf_counter()
    response = await call_next(request)
    duration_ms = int((time.perf_counter() - started_at) * 1000)
    should_record = (
        request.url.path.startswith("/api/")
        and request.url.path not in {
            "/api/session", "/api/hunts/status", "/api/audit/logs",
            "/api/dashboard/operations",
        }
        and (request.method not in {"GET", "HEAD"} or response.status_code >= 400)
    )
    if should_record:
        asyncio.create_task(_record_audit_event({
            "level": "ERROR" if response.status_code >= 500 else (
                "WARNING" if response.status_code >= 400 else "INFO"
            ),
            "service": "thos-ui",
            "category": "api_request",
            "actor": valid_user,
            "action": request.method,
            "resource": request.url.path,
            "status_code": response.status_code,
            "duration_ms": duration_ms,
            "message": (
                f"{request.method} {request.url.path} completed with "
                f"HTTP {response.status_code}"
            ),
            "context": {"query": str(request.url.query)[:1000]},
        }))
    return response


class HuntRequest(BaseModel):
    hypothesis_id: str | None = None
    hypothesis_text: str | None = None
    hypothesis_tactic: str = ""
    hypothesis_technique: str = ""
    siem_type: str = "folder"
    siem_types: list[str] = Field(default_factory=list, max_length=16)
    log_source_path: str | None = None
    max_iterations: int | None = Field(default=None, ge=1, le=5)
    lookback_minutes: int | None = Field(default=None, ge=1, le=525600)
    max_lookback_minutes: int = Field(default=10080, ge=60, le=525600)
    cover_style: str = Field(default="1", pattern="^[12]$")

    @model_validator(mode="after")
    def require_hypothesis(self):
        if not (self.hypothesis_id or "").strip() and not (self.hypothesis_text or "").strip():
            raise ValueError("hypothesis_id or hypothesis_text is required")
        return self


class LogSearchTranslateRequest(BaseModel):
    portable_query: str = Field(min_length=3, max_length=8_000)
    siem_type: str = Field(pattern=r"^[a-z][a-z0-9_-]{0,63}$")
    lookback_minutes: int = Field(default=1440, ge=1, le=525_600)


class LogSearchRunRequest(LogSearchTranslateRequest):
    query: str = Field(min_length=1, max_length=8_000)
    limit: int = Field(default=250, ge=1, le=2_000)
    log_source_path: str | None = Field(default=None, max_length=4_096)


class LogSearchExportRequest(BaseModel):
    portable_query: str = Field(min_length=1, max_length=8_000)
    query: str = Field(min_length=1, max_length=8_000)
    siem_type: str = Field(pattern=r"^[a-z][a-z0-9_-]{0,63}$")
    lookback_minutes: int = Field(default=1440, ge=1, le=525_600)
    executed_at: str = Field(default="", max_length=100)
    field_mapping: dict[str, str] = Field(default_factory=dict)
    logs: list[dict] = Field(default_factory=list, max_length=2_000)


class RiskResolutionRequest(BaseModel):
    note: str = Field(default="", max_length=1000)


class AnomalyLeadUpdateRequest(BaseModel):
    status: str = Field(pattern="^(active|closed|suppressed)$")
    note: str = Field(default="", max_length=1000)


class AnomalyEvaluateRequest(BaseModel):
    source: str = Field(pattern="^(wazuh|logrhythm|splunk|qradar|elasticsearch)$")
    lookback_minutes: int | None = Field(default=None, ge=1, le=1440)


def _excel_safe_value(value):
    if value is None or isinstance(value, (bool, int, float)):
        return value
    if isinstance(value, (dict, list, tuple)):
        value = json.dumps(value, ensure_ascii=False, separators=(",", ":"), default=str)
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", str(value))[:32_767]
    if text.startswith(("=", "+", "-", "@")):
        text = "'" + text
    return text


def _flatten_log_record(record: dict, prefix: str = "") -> dict[str, object]:
    flattened: dict[str, object] = {}
    for key, value in record.items():
        name = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(value, dict):
            flattened.update(_flatten_log_record(value, name))
        else:
            flattened[name] = _excel_safe_value(value)
    return flattened


def _build_log_search_workbook(payload: LogSearchExportRequest) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    logs_sheet = workbook.active
    logs_sheet.title = "Logs"
    flattened = [_flatten_log_record(record) for record in payload.logs]
    priority = [
        "@timestamp", "timestamp", "event_time", "agent.name", "host.name",
        "host", "user.name", "user", "process.name", "process.command_line",
        "source.ip", "src_ip", "destination.ip", "dst_ip", "rule.id",
        "rule.description", "message", "full_log",
    ]
    discovered = list(dict.fromkeys(
        key for record in flattened for key in record
    ))
    columns = [key for key in priority if key in discovered]
    columns.extend(key for key in discovered if key not in columns)
    columns = columns[:200]
    header_fill = PatternFill("solid", fgColor="263B80")
    header_font = Font(color="FFFFFF", bold=True)
    if columns:
        logs_sheet.append(columns)
        for cell in logs_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for record in flattened:
            logs_sheet.append([record.get(column) for column in columns])
        logs_sheet.freeze_panes = "A2"
        logs_sheet.auto_filter.ref = logs_sheet.dimensions
        logs_sheet.row_dimensions[1].height = 30
        for index, column in enumerate(columns, start=1):
            sample = [str(column), *[
                str(record.get(column) or "") for record in flattened[:100]
            ]]
            logs_sheet.column_dimensions[logs_sheet.cell(1, index).column_letter].width = min(
                60, max(12, max(len(value) for value in sample) + 2)
            )
    else:
        logs_sheet.append(["No records returned"])
        logs_sheet["A1"].fill = header_fill
        logs_sheet["A1"].font = header_font
        logs_sheet.column_dimensions["A"].width = 24

    metadata = workbook.create_sheet("Search metadata")
    metadata.append(["Property", "Value"])
    metadata_rows = [
        ("Telemetry source", payload.siem_type),
        ("Portable query", payload.portable_query),
        ("Executed target query", payload.query),
        ("Lookback minutes", payload.lookback_minutes),
        ("Executed at", payload.executed_at),
        ("Exported at", datetime.now(timezone.utc).isoformat()),
        ("Exported records", len(payload.logs)),
    ]
    for row in metadata_rows:
        metadata.append([_excel_safe_value(row[0]), _excel_safe_value(row[1])])
    for cell in metadata[1]:
        cell.fill = header_fill
        cell.font = header_font
    metadata.freeze_panes = "A2"
    metadata.column_dimensions["A"].width = 24
    metadata.column_dimensions["B"].width = 100
    for row in metadata.iter_rows(min_row=2, max_col=2):
        row[1].alignment = Alignment(vertical="top", wrap_text=True)

    mappings = workbook.create_sheet("Field mapping")
    mappings.append(["Normalized field", "Target SIEM field"])
    for normalized, vendor in sorted(payload.field_mapping.items()):
        mappings.append([_excel_safe_value(normalized), _excel_safe_value(vendor)])
    for cell in mappings[1]:
        cell.fill = header_fill
        cell.font = header_font
    mappings.freeze_panes = "A2"
    mappings.auto_filter.ref = mappings.dimensions
    mappings.column_dimensions["A"].width = 28
    mappings.column_dimensions["B"].width = 70

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


class LoginRequest(BaseModel):
    username: str = Field(min_length=1, max_length=128)
    password: str = Field(min_length=1, max_length=512)


async def _upstream_json(
    method: str,
    path: str,
    *,
    upstream_timeout: httpx.Timeout | None = None,
    **kwargs,
):
    timeout = upstream_timeout or httpx.Timeout(60.0, connect=10.0)
    try:
        async with httpx.AsyncClient(timeout=timeout) as client:
            response = await client.request(
                method,
                f"{ORCHESTRATOR_URL}{path}",
                headers=_UPSTREAM_HEADERS,
                **kwargs,
            )
    except httpx.TimeoutException as exc:
        raise HTTPException(
            status_code=504,
            detail=f"Orchestrator request timed out for {path}",
        ) from exc
    except httpx.RequestError as exc:
        raise HTTPException(
            status_code=502,
            detail=f"Could not reach the orchestrator for {path}",
        ) from exc
    if response.is_error:
        try:
            detail = response.json().get("detail", response.text)
        except ValueError:
            detail = response.text
        raise HTTPException(status_code=response.status_code, detail=detail or "orchestrator request failed")
    return response.json()


@app.get("/health")
async def health():
    return {"status": "ok"}


@app.post("/api/auth/login")
async def login(credentials: LoginRequest):
    username = _valid_account(credentials.username, credentials.password)
    if not username:
        asyncio.create_task(_record_audit_event({
            "level": "WARNING",
            "service": "thos-ui",
            "category": "authentication",
            "actor": credentials.username[:160],
            "action": "login_failed",
            "resource": "/api/auth/login",
            "status_code": 401,
            "message": "Interactive sign-in was rejected",
            "context": {},
        }))
        raise HTTPException(status_code=401, detail="Invalid username or password")
    user = control_plane.get_user(username) or {}
    response = JSONResponse({
        "analyst": username,
        **control_plane.public_user(user),
        "branding": control_plane.public_branding(),
    })
    response.set_cookie(
        SESSION_COOKIE,
        _session_token(username),
        max_age=SESSION_TTL_SECONDS,
        httponly=True,
        secure=SESSION_SECURE_COOKIE,
        samesite="strict",
        path="/",
    )
    asyncio.create_task(_record_audit_event({
        "level": "INFO",
        "service": "thos-ui",
        "category": "authentication",
        "actor": username,
        "action": "login",
        "resource": "/api/auth/login",
        "status_code": 200,
        "message": "Interactive sign-in succeeded",
        "context": {"role": user.get("role", "Expert")},
    }))
    return response


@app.post("/api/auth/logout")
async def logout(request: Request):
    username = _session_user(request.cookies.get(SESSION_COOKIE, ""))
    response = JSONResponse({"status": "signed_out"})
    response.delete_cookie(SESSION_COOKIE, path="/", samesite="strict")
    if username:
        asyncio.create_task(_record_audit_event({
            "level": "INFO",
            "service": "thos-ui",
            "category": "authentication",
            "actor": username,
            "action": "logout",
            "resource": "/api/auth/logout",
            "status_code": 200,
            "message": "Interactive session signed out",
            "context": {},
        }))
    return response


@app.get("/api/session")
async def session(request: Request):
    user = control_plane.get_user(request.state.analyst) or {}
    return {
        "analyst": request.state.analyst,
        **control_plane.public_user(user),
        "permissions": sorted(request.state.permissions),
        "branding": control_plane.public_branding(),
    }


@app.get("/api/dashboard/operations")
async def dashboard_operations(request: Request, hours: int = 24):
    bounded_hours = max(1, min(hours, 24 * 365))
    payload = await _upstream_json(
        "GET", "/dashboard/operations", params={"hours": bounded_hours}
    )
    config = read_config()
    schedules = [
        *config.get("hypothesis_schedules", []),
        *(config.get("sigma", {}).get("schedules", []) or []),
        *(config.get("yara", {}).get("schedules", []) or []),
    ]
    sources = control_plane.telemetry_sources(config)
    integrations = config.get("integrations", {}) or {}
    cutoff = datetime.now(timezone.utc).timestamp() - bounded_hours * 3600
    report_count = sum(
        path.is_file() and path.stat().st_mtime >= cutoff
        for path in REPORTS_DIR.glob("*.md")
    ) if REPORTS_DIR.exists() else 0
    payload["platform"] = {
        "telemetry_sources": len(sources.get("items") or []),
        "enabled_schedules": sum(bool(item.get("enabled", True)) for item in schedules),
        "schedule_failures": sum(
            str(item.get("last_status") or "") == "failed" for item in schedules
        ),
        "connected_integrations": sum(
            isinstance(value, dict) and value.get("connection_status") == "connected"
            for value in integrations.values()
        ),
        "report_library_created": report_count,
    }
    return payload


@app.get("/api/risks")
async def actionable_risks(
    request: Request,
    limit: int = 500,
    hours: int = 0,
    refresh: bool = False,
):
    control_plane.require_feature(request, "reports")
    return await _upstream_json(
        "GET",
        "/risks",
        upstream_timeout=httpx.Timeout(300.0, connect=10.0),
        params={
            "limit": max(1, min(limit, 2000)),
            "hours": max(1, min(hours, 24 * 365 * 10)) if hours else 0,
            "refresh": bool(refresh),
        }
    )


@app.patch("/api/risks/{risk_id}/resolve")
async def resolve_actionable_risk(
    risk_id: str,
    payload: RiskResolutionRequest,
    request: Request,
):
    control_plane.require_feature(request, "reports")
    if request.state.role not in {"Admin", "SME"}:
        raise HTTPException(
            status_code=403,
            detail="Only Admin and SME users can resolve risks",
        )
    if not re.fullmatch(r"[A-Za-z0-9._:-]{1,160}", risk_id):
        raise HTTPException(status_code=422, detail="Invalid risk identifier")
    return await _upstream_json(
        "PATCH",
        f"/risks/{risk_id}/resolve",
        json={
            "actor": request.state.analyst,
            "role": request.state.role,
            "note": payload.note,
        },
    )


@app.get("/api/anomaly-leads")
async def anomaly_leads(
    request: Request,
    status: str = "active",
    limit: int = 100,
    entity_type: str = "",
    entity_name: str = "",
):
    control_plane.require_feature(request, "hunts")
    return await _upstream_json(
        "GET",
        "/anomalies/leads",
        params={
            "status": status,
            "limit": max(1, min(limit, 1000)),
            "entity_type": entity_type[:40],
            "entity_name": entity_name[:240],
        },
    )


@app.get("/api/anomaly-status")
async def anomaly_monitor_status(request: Request):
    control_plane.require_feature(request, "hunts")
    payload = await _upstream_json("GET", "/anomalies/status")
    settings = read_config().get("anomaly_monitoring", {}) or {}
    return {
        **payload,
        "enabled": bool(settings.get("enabled", True)),
        "interval_minutes": int(settings.get("interval_minutes", 15)),
        "minimum_baseline_buckets": int(settings.get("minimum_baseline_buckets", 24)),
        "scheduler": {
            "status": settings.get("last_status", "never"),
            "last_started_at": settings.get("last_started_at", ""),
            "last_completed_at": settings.get("last_completed_at", ""),
            "last_error": settings.get("last_error", ""),
            "results": settings.get("last_results", []),
        },
    }


@app.post("/api/anomaly-leads/evaluate")
async def evaluate_anomaly_leads(payload: AnomalyEvaluateRequest, request: Request):
    control_plane.require_feature(request, "hunts")
    if request.state.role not in {"Admin", "SME"}:
        raise HTTPException(status_code=403, detail="Only Admin and SME users can run anomaly evaluation")
    if not control_plane.is_active_telemetry_source(payload.source):
        raise HTTPException(status_code=422, detail="The selected telemetry source is not connected")
    return await _upstream_json(
        "POST",
        "/anomalies/evaluate",
        upstream_timeout=httpx.Timeout(180.0, connect=10.0),
        json=payload.model_dump(exclude_none=True),
    )


@app.patch("/api/anomaly-leads/{lead_id}")
async def update_anomaly_lead(
    lead_id: str,
    payload: AnomalyLeadUpdateRequest,
    request: Request,
):
    control_plane.require_feature(request, "hunts")
    if request.state.role not in {"Admin", "SME"}:
        raise HTTPException(status_code=403, detail="Only Admin and SME users can manage anomaly leads")
    if not re.fullmatch(r"ANOM-[A-F0-9]{20}", lead_id):
        raise HTTPException(status_code=422, detail="Invalid anomaly lead identifier")
    return await _upstream_json(
        "PATCH",
        f"/anomalies/leads/{lead_id}",
        json={
            "status": payload.status,
            "actor": request.state.analyst,
            "note": payload.note,
        },
    )


@app.post("/api/anomaly-leads/{lead_id}/link/{hunt_id}")
async def link_anomaly_lead_hunt(lead_id: str, hunt_id: str, request: Request):
    control_plane.require_feature(request, "hunts")
    if not re.fullmatch(r"ANOM-[A-F0-9]{20}", lead_id):
        raise HTTPException(status_code=422, detail="Invalid anomaly lead identifier")
    if not re.fullmatch(r"[0-9a-fA-F-]{36}", hunt_id):
        raise HTTPException(status_code=422, detail="Invalid hunt identifier")
    return await _upstream_json(
        "POST",
        f"/anomalies/leads/{lead_id}/link",
        json={"hunt_id": hunt_id},
    )


def _require_active_log_source(request: Request, siem_type: str) -> None:
    control_plane.require_feature(request, "hunts")
    if not control_plane.is_active_telemetry_source(siem_type):
        raise HTTPException(
            status_code=422,
            detail="Select a connected telemetry source before running log search",
        )


@app.get("/api/log-search/context/{siem_type}")
async def log_search_context(siem_type: str, request: Request):
    _require_active_log_source(request, siem_type)
    return await _upstream_json("GET", f"/log-search/context/{siem_type}")


@app.post("/api/log-search/translate")
async def translate_log_search(payload: LogSearchTranslateRequest, request: Request):
    _require_active_log_source(request, payload.siem_type)
    return await _upstream_json(
        "POST",
        "/log-search/translate",
        upstream_timeout=httpx.Timeout(300.0, connect=10.0),
        json=payload.model_dump(),
    )


@app.post("/api/log-search/run")
async def run_log_search(payload: LogSearchRunRequest, request: Request):
    _require_active_log_source(request, payload.siem_type)
    return await _upstream_json(
        "POST",
        "/log-search/run",
        upstream_timeout=httpx.Timeout(300.0, connect=10.0),
        json=payload.model_dump(),
    )


@app.post("/api/log-search/export")
async def export_log_search(payload: LogSearchExportRequest, request: Request):
    _require_active_log_source(request, payload.siem_type)
    workbook = await asyncio.to_thread(_build_log_search_workbook, payload)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    return StreamingResponse(
        BytesIO(workbook),
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                f'attachment; filename="thos-log-search-{payload.siem_type}-{stamp}.xlsx"'
            ),
            "X-Content-Type-Options": "nosniff",
        },
    )


@app.get("/api/audit/logs")
async def audit_logs(
    request: Request,
    hours: int = 24,
    limit: int = 500,
    level: str = "all",
    query: str = "",
):
    control_plane.require_feature(request, "settings", sme_only=True)
    return await _upstream_json(
        "GET",
        "/audit/logs",
        params={
            "hours": max(1, min(hours, 24 * 365)),
            "limit": max(1, min(limit, 2000)),
            "level": level,
            "query": query[:500],
        },
    )


@app.get("/api/hypotheses")
async def hypotheses(request: Request):
    control_plane.require_feature(request, "hunts")
    hearth, last_runs = await asyncio.gather(
        _upstream_json("GET", "/hypotheses"),
        _upstream_json("GET", "/hypotheses/last-runs"),
    )
    recent = {item.get("hypothesis_id"): item for item in last_runs}
    custom = read_config().get("custom_hypotheses", [])
    return [
        {**normalize_hypothesis(item),
         "severity": control_plane.hypothesis_severity(item), **({
            "last_ran_at": recent[item.get("id")].get("last_ran_at"),
            "last_run_status": recent[item.get("id")].get("status"),
        } if item.get("id") in recent else {})}
        for item in [*(hearth if isinstance(hearth, list) else []), *custom]
    ]


@app.get("/api/hunts/status")
async def hunt_status(request: Request):
    control_plane.require_feature(request, "hunts")
    try:
        return await _upstream_json("GET", "/hunt/status")
    except httpx.RequestError:
        # The UI polls during rolling restarts; an unavailable Orchestrator is
        # not an application error and must not flood the gateway logs.
        return {"active": False, "active_count": 0, "available": False}


@app.get("/api/hunts/history")
async def hunt_history(request: Request, limit: int = 100):
    control_plane.require_feature(request, "reports")
    return await _upstream_json("GET", f"/hunts?limit={max(1, min(limit, 500))}")


@app.delete("/api/hunts/history")
async def clear_hunt_history(request: Request):
    control_plane.require_feature(request, "reports", admin_only=True)
    return await _upstream_json("DELETE", "/hunts")


@app.get("/api/hunts/{hunt_id}/progress")
async def hunt_progress(hunt_id: str, request: Request):
    control_plane.require_feature(request, "hunts")
    return await _upstream_json("GET", f"/hunts/{hunt_id}/progress")


@app.post("/api/hunts/stream")
async def stream_hunt(hunt: HuntRequest, request: Request):
    control_plane.require_feature(request, "hunts")
    sources = list(dict.fromkeys(
        source.strip().lower()
        for source in ([hunt.siem_type, *hunt.siem_types])
        if source.strip()
    ))
    inactive = [
        source for source in sources
        if not control_plane.is_active_telemetry_source(source)
    ]
    if inactive:
        raise HTTPException(
            status_code=422,
            detail=(
                "Telemetry source is not active; save and successfully test "
                f"it in Configuration: {', '.join(inactive)}"
            ),
        )
    status = await _upstream_json("GET", "/hunt/status")
    if status.get("active"):
        raise HTTPException(status_code=409, detail="A hunt is already running. Wait for it to complete before starting another hypothesis.")
    payload = hunt.model_dump()
    if payload.get("max_iterations") is None:
        payload["max_iterations"] = max(1, min(5, int(get_value("general", "default_iterations", default=2))))
    payload["siem_types"] = sources
    payload["hunter_name"] = request.state.analyst

    async def relay() -> AsyncIterator[bytes]:
        timeout = httpx.Timeout(connect=10.0, read=900.0, write=30.0, pool=10.0)
        try:
            async with httpx.AsyncClient(timeout=timeout) as client:
                async with client.stream(
                    "POST",
                    f"{ORCHESTRATOR_URL}/hunt/stream",
                    headers=_UPSTREAM_HEADERS,
                    json=payload,
                ) as response:
                    if response.is_error:
                        body = await response.aread()
                        yield (json.dumps({
                            "event": "error",
                            "error": body.decode("utf-8", errors="replace") or f"orchestrator returned {response.status_code}",
                        }) + "\n").encode()
                        return
                    async for chunk in response.aiter_bytes():
                        yield chunk
        except Exception as exc:  # noqa: BLE001 - converted to a stream event
            logger.exception("hunt stream proxy failed")
            yield (json.dumps({"event": "error", "error": str(exc)}) + "\n").encode()

    return StreamingResponse(relay(), media_type="application/x-ndjson")


def _safe_report(filename: str) -> Path:
    if not filename or Path(filename).name != filename or not filename.lower().endswith(".md"):
        raise HTTPException(status_code=404, detail="report not found")
    root = REPORTS_DIR.resolve()
    candidate = (root / filename).resolve()
    if candidate.parent != root or not candidate.is_file():
        raise HTTPException(status_code=404, detail="report not found")
    return candidate


def _first_heading(markdown: str, fallback: str) -> str:
    for line in markdown.splitlines():
        match = re.match(r"^#\s+(.+?)\s*$", line)
        if match:
            return re.sub(r"[*_`]", "", match.group(1)).strip()
    return fallback


def _hunt_id(markdown: str) -> str:
    patterns = (
        # Current structured reports render identifiers in a blockquoted table.
        # Keep the parser format-aware rather than deriving IDs from filenames so
        # renamed/imported reports continue to associate with their audit row.
        r"(?mi)^\s*>?\s*\|\s*Hunt ID\s*\|\s*`?([0-9a-fA-F-]{20,})",
        r"\*\*Hunt ID:?\*\*\s*[:|]?\s*`?([0-9a-fA-F-]{20,})",
        r"Hunt\s+`([0-9a-fA-F-]{20,})`",
    )
    for pattern in patterns:
        match = re.search(pattern, markdown)
        if match:
            return match.group(1)
    return ""


def _report_metadata(path: Path, content: str | None = None) -> dict:
    text = content if content is not None else path.read_text(encoding="utf-8", errors="replace")
    text = _clean_report_presentation(text)
    stat = path.stat()
    report_type = "forensic" if (
        path.name.upper().startswith("FORENSIC_")
        or "Report classification:** Digital forensic technical report" in text
    ) else "hunt"
    case_match = re.search(r"\*\*Case ID:\*\*\s*`([^`]+)`", text) if report_type == "forensic" else None
    return {
        "filename": path.name,
        "title": _first_heading(text, path.stem.replace("_", " ")),
        "type": report_type,
        "hunt_id": _hunt_id(text),
        "case_id": case_match.group(1) if case_match else "",
        "modified": datetime.fromtimestamp(stat.st_mtime, timezone.utc).isoformat(),
        "size": stat.st_size,
    }


def _clean_report_presentation(markdown: str) -> str:
    """Apply current report presentation policy to new and historical files."""
    cleaned = re.sub(
        r"(?im)^\*Generated by THOS \([^)]*\).*?\*\s*$",
        "*Generated by THOS (On-Prem AI Threat Hunting Operating System).*",
        markdown,
    )
    cleaned = re.sub(
        r"(?im)^\*This report was produced by an AI reasoning pipeline.*?human analyst should validate findings before action\.\*\s*$",
        "*A human analyst should validate findings before action.*",
        cleaned,
    )
    cleaned = re.sub(r"(?i)\bSigmaHQ\b", "Community", cleaned)
    cleaned = re.sub(r"(?i)\bpySigma\b", "audited rule compiler", cleaned)
    cleaned = re.sub(r"(?i)\bSigma\b", "detection rule", cleaned)
    cleaned = re.sub(
        r"(?i)\bhuman approval\b|\bmanual approval\b",
        "analyst review",
        cleaned,
    )
    cleaned = re.sub(
        r"(?i)\bpending approvals?\b",
        "pending analyst reviews",
        cleaned,
    )
    # Historical model output occasionally labelled an absence claim as hard
    # evidence. Preserve the finding text, but correct the evidence class in
    # previews and downloads; absence is circumstantial unless a deterministic
    # coverage proof establishes it.
    cleaned = re.sub(
        r"(?im)^(\s*-\s*)\[\s*hard-evidence\s*\]"
        r"(?=[^\n]*(?:no evidence|not observed|not present|"
        r"no [^.\n]{0,80} (?:found|detected|identified)|absence of))",
        r"\1[circumstantial]",
        cleaned,
    )
    cleaned = re.sub(
        "["
        "\U0001F1E0-\U0001F1FF"
        "\U0001F300-\U0001FAFF"
        "\u2600-\u27BF"
        "\uFE0F\u200D"
        "]",
        "",
        cleaned,
    )
    cleaned = re.sub(r"(?m)^(#{1,6})[ \t]{2,}", r"\1 ", cleaned)
    cleaned = re.sub(r"\[\s+([^\]]+)\]", r"[\1]", cleaned)
    return re.sub(r"[🚀🔎📊🔍🧠✅⚠️🛡️📋📌📄📈📉🔐🧾]", "", cleaned)


@app.get("/api/reports")
async def list_reports(request: Request):
    control_plane.require_feature(request, "reports")
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    reports = [_report_metadata(path) for path in REPORTS_DIR.glob("*.md") if path.is_file()]
    return sorted(reports, key=lambda item: item["modified"], reverse=True)


@app.delete("/api/reports/{filename}")
async def delete_report(filename: str, request: Request):
    control_plane.require_feature(request, "reports", admin_only=True)
    path = _safe_report(filename)
    trash = (REPORTS_DIR / ".trash").resolve()
    if trash.parent != REPORTS_DIR.resolve():
        raise HTTPException(status_code=500, detail="invalid report archive path")
    trash.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%fZ")
    archived = (trash / f"{stamp}_{path.name}").resolve()
    if archived.parent != trash:
        raise HTTPException(status_code=500, detail="invalid report archive target")
    path.replace(archived)
    logger.info(
        "report removed from the active library",
        extra={"report": path.name, "actor": request.state.analyst, "archived": str(archived)},
    )
    return {"deleted": path.name, "recoverable": True, "archived_path": str(archived)}


def _safe_evidence_name(filename: str, fallback: str) -> str:
    base = Path(filename or "").name
    safe = re.sub(r"[^A-Za-z0-9._() -]+", "_", base).strip(" .")
    return safe[:180] or fallback


async def _new_forensic_case_dir(label: str) -> Path:
    date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    prefix = datetime.now(timezone.utc).strftime("%Y%m%d")
    safe_label = re.sub(r"[^A-Za-z0-9_-]+", "-", label).strip("-")[:80] or "evidence"
    root = FORENSIC_ROOT.resolve()
    day_root = (root / date).resolve()
    if day_root.parent != root:
        raise HTTPException(status_code=500, detail="invalid forensic storage path")
    day_root.mkdir(parents=True, exist_ok=True)
    async with _forensic_intake_lock:
        for serial in range(1, 10000):
            candidate = (day_root / f"{prefix}-{serial:04d}-{safe_label}").resolve()
            if candidate.parent != day_root:
                continue
            try:
                candidate.mkdir()
                return candidate
            except FileExistsError:
                continue
    raise HTTPException(status_code=507, detail="daily forensic case serial space is exhausted")


def _yara_scan_summary(payload: dict) -> dict:
    scan = payload.get("scan_result") or {}
    return {
        "scan_id": payload.get("scan_id", ""),
        "scan_title": payload.get("scan_title", ""),
        "artifact_type": payload.get("artifact_type", ""),
        "original_name": payload.get("original_name", ""),
        "size_bytes": payload.get("size_bytes", 0),
        "sha256": payload.get("sha256", ""),
        "examiner": payload.get("examiner", ""),
        "received_at": payload.get("received_at", ""),
        "completed_at": payload.get("completed_at", ""),
        "status": payload.get("status", "unknown"),
        "match_count": scan.get("match_count", 0),
        "static_finding_count": int(
            (payload.get("static_analysis") or {}).get("finding_count", 0) or 0
        ),
        "duration_ms": scan.get("duration_ms", 0),
        "error": payload.get("error", ""),
    }


def _yara_scan_manifests() -> list[tuple[float, Path, dict]]:
    root = FORENSIC_ROOT.resolve()
    if not root.is_dir():
        return []
    manifests = []
    for path in root.rglob("_thos_yara_scan.json"):
        try:
            resolved = path.resolve()
            if root not in resolved.parents:
                continue
            payload = json.loads(resolved.read_text(encoding="utf-8"))
            if not isinstance(payload, dict) or not payload.get("scan_id"):
                continue
            manifests.append((resolved.stat().st_mtime, resolved, payload))
        except (OSError, ValueError, TypeError):
            logger.warning("could not read forensic YARA scan manifest", extra={"path": str(path)})
    return sorted(manifests, key=lambda item: item[0], reverse=True)


@app.post("/api/forensics/yara-scans")
async def create_yara_forensic_scan(
    request: Request,
    sample: UploadFile = File(...),
    scan_title: str = Form(...),
    artifact_type: str = Form("suspicious_file"),
    acquired_from: str = Form(""),
    notes: str = Form(""),
):
    control_plane.require_feature(request, "forensics")
    title = scan_title.strip()
    if not title or len(title) > 300:
        raise HTTPException(status_code=422, detail="scan title must contain 1-300 characters")
    if artifact_type not in {"memory_dump", "process_dump", "suspicious_file"}:
        raise HTTPException(
            status_code=422,
            detail="artifact type must be memory_dump, process_dump, or suspicious_file",
        )

    scan_id = str(uuid.uuid4())
    case_dir = await _new_forensic_case_dir(f"yara-{title}")
    original_name = _safe_evidence_name(sample.filename or "", "suspicious-artifact.bin")
    target = (case_dir / f"Y0001_{original_name}").resolve()
    if target.parent != case_dir:
        raise HTTPException(status_code=422, detail="invalid dump filename")

    digest = hashlib.sha256()
    size = 0
    received_at = datetime.now(timezone.utc).isoformat()
    manifest_path = case_dir / "_thos_yara_scan.json"
    base_manifest = {
        "scan_id": scan_id,
        "scan_title": title,
        "artifact_type": artifact_type,
        "original_name": original_name,
        "stored_name": target.name,
        "examiner": request.state.analyst,
        "received_at": received_at,
        "acquired_from": acquired_from.strip()[:1000],
        "notes": notes.strip()[:10000],
        "agent": {
            "agent_id": "file_memory_static_analysis",
            "agent_name": "File and Memory Static Analysis Agent",
            "activity": (
                "Runs content-routed, non-executing forensic triage and the enabled "
                "actionable YARA rule bundle against the preserved artifact."
            ),
        },
    }
    try:
        with target.open("xb") as handle:
            while chunk := await sample.read(1024 * 1024):
                size += len(chunk)
                if size > FORENSIC_MAX_FILE_BYTES:
                    raise HTTPException(
                        status_code=413,
                        detail=f"{original_name} exceeds the forensic artifact upload limit",
                    )
                digest.update(chunk)
                handle.write(chunk)
            handle.flush()
            os.fsync(handle.fileno())
        try:
            target.chmod(0o440)
        except OSError:
            logger.warning("could not mark memory dump read-only", extra={"scan_id": scan_id})

        sha256 = digest.hexdigest()
        pending = {
            **base_manifest,
            "size_bytes": size,
            "sha256": sha256,
            "status": "scanning",
        }
        manifest_path.write_text(json.dumps(pending, indent=2), encoding="utf-8")
        static_analysis = await _upstream_json(
            "POST",
            "/forensics/static-scan",
            json={
                "path": str(target),
                "sha256": sha256,
                "artifact_type": artifact_type,
            },
            timeout=3600.0,
        )
        result = await _upstream_json(
            "POST",
            "/yara/scan",
            json={
                "path": str(target),
                "recursive": False,
                "analysis_profile": (
                    "memory"
                    if artifact_type in {"memory_dump", "process_dump"}
                    else "evidence"
                ),
            },
            timeout=1200.0,
        )
        file_result = (result.get("results") or [{}])[0]
        status = (
            "failed"
            if result.get("errors") or file_result.get("status") == "skipped"
            else "matched"
            if result.get("match_count") or static_analysis.get("finding_count")
            else "clean"
        )
        completed = {
            **pending,
            "status": status,
            "completed_at": datetime.now(timezone.utc).isoformat(),
            "scan_result": result,
            "static_analysis": static_analysis,
        }
        manifest_path.write_text(json.dumps(completed, indent=2, default=str), encoding="utf-8")
        logger.info(
            "forensic file and memory rule scan completed",
            extra={
                "scan_id": scan_id,
                "artifact_type": artifact_type,
                "size_bytes": size,
                "match_count": result.get("match_count", 0),
                "status": status,
            },
        )
        return completed
    except Exception as exc:
        failed = {
            **base_manifest,
            "size_bytes": size,
            "sha256": digest.hexdigest(),
            "status": "failed",
            "completed_at": datetime.now(timezone.utc).isoformat(),
            "error": str(exc),
        }
        try:
            manifest_path.write_text(json.dumps(failed, indent=2), encoding="utf-8")
        except OSError:
            pass
        logger.exception("forensic file and memory rule scan failed", extra={"scan_id": scan_id})
        raise
    finally:
        await sample.close()


@app.get("/api/forensics/yara-scans")
async def list_yara_forensic_scans(request: Request, limit: int = 100):
    control_plane.require_feature(request, "forensics")
    bounded = max(1, min(limit, 500))
    return [
        _yara_scan_summary(payload)
        for _, _, payload in _yara_scan_manifests()[:bounded]
    ]


@app.get("/api/forensics/yara-scans/{scan_id}")
async def read_yara_forensic_scan(scan_id: str, request: Request):
    control_plane.require_feature(request, "forensics")
    if not re.fullmatch(r"[0-9a-fA-F-]{36}", scan_id):
        raise HTTPException(status_code=422, detail="invalid forensic scan identifier")
    for _, _, payload in _yara_scan_manifests():
        if payload.get("scan_id") == scan_id:
            return payload
    raise HTTPException(status_code=404, detail="forensic YARA scan not found")


@app.post("/api/forensics/cases", status_code=202)
async def create_forensic_case(
    request: Request,
    files: list[UploadFile] = File(...),
    case_title: str = Form(...),
    acquired_from: str = Form(""),
    legal_authority: str = Form(""),
    notes: str = Form(""),
):
    control_plane.require_feature(request, "forensics")
    title = case_title.strip()
    if not title or len(title) > 300:
        raise HTTPException(status_code=422, detail="case title must contain 1-300 characters")
    if not files:
        raise HTTPException(status_code=422, detail="at least one evidence file is required")
    case_id = str(uuid.uuid4())
    case_dir = await _new_forensic_case_dir(title)
    evidence, case_bytes = [], 0
    try:
        for index, upload in enumerate(files, start=1):
            evidence_id = f"E{index:04d}"
            original_name = _safe_evidence_name(upload.filename or "", f"evidence-{index}")
            stored_name = f"{evidence_id}_{original_name}"
            target = (case_dir / stored_name).resolve()
            if target.parent != case_dir:
                raise HTTPException(status_code=422, detail="invalid evidence filename")
            digest = hashlib.sha256()
            size = 0
            with target.open("xb") as handle:
                while chunk := await upload.read(1024 * 1024):
                    size += len(chunk)
                    case_bytes += len(chunk)
                    if size > FORENSIC_MAX_FILE_BYTES:
                        raise HTTPException(status_code=413, detail=f"{original_name} exceeds the evidence file limit")
                    if case_bytes > FORENSIC_MAX_CASE_BYTES:
                        raise HTTPException(status_code=413, detail="forensic case exceeds the total upload limit")
                    digest.update(chunk)
                    handle.write(chunk)
                handle.flush()
                os.fsync(handle.fileno())
            evidence.append({
                "evidence_id": evidence_id,
                "original_name": original_name,
                "stored_name": stored_name,
                "size_bytes": size,
                "sha256": digest.hexdigest(),
                "content_type": upload.content_type or "application/octet-stream",
            })
        manifest = {
            "case_id": case_id,
            "case_title": title,
            "examiner": request.state.analyst,
            "received_at": datetime.now(timezone.utc).isoformat(),
            "acquired_from": acquired_from.strip()[:1000],
            "legal_authority": legal_authority.strip()[:2000],
            "notes": notes.strip()[:10000],
            "evidence": evidence,
        }
        manifest_path = case_dir / "_thos_chain_of_custody.json"
        manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
        for item in evidence:
            try:
                (case_dir / item["stored_name"]).chmod(0o440)
            except OSError:
                logger.warning("could not mark evidence read-only", extra={"evidence": item["evidence_id"]})
        result = await _upstream_json("POST", "/forensics/analyze", json={
            "case_id": case_id,
            "case_title": title,
            "case_dir": str(case_dir),
            "examiner": request.state.analyst,
        })
        return {**result, "evidence": evidence}
    except Exception as exc:
        try:
            (case_dir / "_thos_intake_failed.json").write_text(json.dumps({
                "case_id": case_id,
                "case_title": title,
                "examiner": request.state.analyst,
                "failed_at": datetime.now(timezone.utc).isoformat(),
                "error": str(exc),
                "completed_evidence": evidence,
                "partial_bytes_received": case_bytes,
            }, indent=2), encoding="utf-8")
        except OSError:
            pass
        logger.exception(
            "forensic intake failed; preserving any uploaded bytes for administrator review",
            extra={"case_id": case_id, "case_dir": str(case_dir)},
        )
        raise
    finally:
        for upload in files:
            await upload.close()


@app.get("/api/forensics")
async def list_forensic_cases(request: Request, limit: int = 100):
    control_plane.require_feature(request, "forensics")
    return await _upstream_json("GET", f"/forensics?limit={max(1, min(limit, 500))}")


@app.get("/api/forensics/tools")
async def read_forensic_tools(request: Request):
    control_plane.require_feature(request, "forensics")
    return await _upstream_json("GET", "/forensics/tools")


@app.get("/api/forensics/{case_id}")
async def read_forensic_case(case_id: str, request: Request):
    control_plane.require_feature(request, "forensics")
    if not re.fullmatch(r"[0-9a-fA-F-]{36}", case_id):
        raise HTTPException(status_code=422, detail="invalid forensic case identifier")
    return await _upstream_json("GET", f"/forensics/{case_id}")


@app.get("/api/reports/{filename}")
async def read_report(filename: str, request: Request):
    control_plane.require_feature(request, "reports")
    path = _safe_report(filename)
    content = _clean_report_presentation(
        path.read_text(encoding="utf-8", errors="replace")
    )
    return {**_report_metadata(path, content), "content": content}


@app.get("/api/reports/{filename}/markdown")
async def download_markdown(filename: str, request: Request):
    control_plane.require_feature(request, "reports")
    path = _safe_report(filename)
    content = _clean_report_presentation(
        path.read_text(encoding="utf-8", errors="replace")
    )
    return Response(
        content=content.encode("utf-8"),
        media_type="text/markdown; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{path.name}"'},
    )


def _register_fonts() -> tuple[str, str]:
    regular_path = Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf")
    bold_path = Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf")
    mono_path = Path("/usr/share/fonts/truetype/dejavu/DejaVuSansMono.ttf")
    if regular_path.is_file() and bold_path.is_file():
        if "THOSDejaVu" not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(TTFont("THOSDejaVu", str(regular_path)))
            pdfmetrics.registerFont(TTFont("THOSDejaVu-Bold", str(bold_path)))
            pdfmetrics.registerFontFamily("THOSDejaVu", normal="THOSDejaVu", bold="THOSDejaVu-Bold")
            if mono_path.is_file():
                pdfmetrics.registerFont(TTFont("THOSMono", str(mono_path)))
        return "THOSDejaVu", "THOSMono" if mono_path.is_file() else "Courier"
    return "Helvetica", "Courier"


def _inline_markdown(value: str) -> str:
    text = escape(value.strip())
    text = re.sub(r"`([^`]+)`", r'<font name="THOSMono" color="#4338ca">\1</font>', text)
    text = re.sub(r"\*\*([^*]+)\*\*", r"<b>\1</b>", text)
    text = re.sub(r"__([^_]+)__", r"<b>\1</b>", text)
    text = re.sub(r"\*([^*]+)\*", r"<i>\1</i>", text)
    text = re.sub(r"\[([^]]+)]\((https?://[^)]+)\)", r'<a href="\2" color="#4f46e5">\1</a>', text)
    return text


def _pdf_styles(font_name: str, mono_name: str):
    sample = getSampleStyleSheet()
    return {
        "title": ParagraphStyle("Title", parent=sample["Title"], fontName=font_name, fontSize=22, leading=28, textColor=colors.HexColor("#101828"), alignment=TA_LEFT, spaceAfter=14),
        "h2": ParagraphStyle("H2", parent=sample["Heading2"], fontName=font_name, fontSize=15, leading=20, textColor=colors.HexColor("#172033"), spaceBefore=16, spaceAfter=8, borderColor=colors.HexColor("#dfe4ec"), borderWidth=0, borderPadding=(0, 0, 5, 0)),
        "h3": ParagraphStyle("H3", parent=sample["Heading3"], fontName=font_name, fontSize=11.5, leading=16, textColor=colors.HexColor("#25324a"), spaceBefore=11, spaceAfter=6),
        "body": ParagraphStyle("Body", parent=sample["BodyText"], fontName=font_name, fontSize=8.8, leading=13.5, textColor=colors.HexColor("#344054"), spaceAfter=7),
        "bullet": ParagraphStyle("Bullet", parent=sample["BodyText"], fontName=font_name, fontSize=8.6, leading=13, leftIndent=12, firstLineIndent=-8, bulletIndent=0, textColor=colors.HexColor("#344054"), spaceAfter=4),
        "quote": ParagraphStyle("Quote", parent=sample["BodyText"], fontName=font_name, fontSize=8.6, leading=13, leftIndent=12, rightIndent=8, borderColor=colors.HexColor("#818cf8"), borderWidth=2, borderPadding=7, backColor=colors.HexColor("#f5f7ff"), textColor=colors.HexColor("#475467"), spaceBefore=5, spaceAfter=8),
        "code": ParagraphStyle("Code", parent=sample["Code"], fontName=mono_name, fontSize=6.8, leading=9.5, leftIndent=7, rightIndent=7, borderPadding=8, backColor=colors.HexColor("#111827"), textColor=colors.HexColor("#e5e7eb"), spaceBefore=6, spaceAfter=9),
        "cell": ParagraphStyle("Cell", parent=sample["BodyText"], fontName=font_name, fontSize=6.8, leading=9.5, textColor=colors.HexColor("#344054")),
        "cell_head": ParagraphStyle("CellHead", parent=sample["BodyText"], fontName=font_name, fontSize=6.8, leading=9.5, textColor=colors.HexColor("#172033")),
    }


_TABLE_SEPARATOR = re.compile(r"^\s*\|?\s*:?-{3,}:?\s*(?:\|\s*:?-{3,}:?\s*)+\|?\s*$")


def _table_cells(line: str) -> list[str]:
    return [cell.strip() for cell in line.strip().strip("|").split("|")]


def _markdown_story(markdown: str, styles: dict) -> list:
    lines = markdown.replace("\r\n", "\n").split("\n")
    story: list = []
    paragraph: list[str] = []
    in_code = False
    code_lines: list[str] = []

    def flush_paragraph() -> None:
        if paragraph:
            story.append(Paragraph(_inline_markdown(" ".join(part.strip() for part in paragraph)), styles["body"]))
            paragraph.clear()

    index = 0
    while index < len(lines):
        line = lines[index]
        stripped = line.strip()
        if stripped.startswith("```"):
            flush_paragraph()
            if in_code:
                story.append(Preformatted("\n".join(code_lines), styles["code"], maxLineLength=108))
                code_lines.clear()
                in_code = False
            else:
                in_code = True
            index += 1
            continue
        if in_code:
            code_lines.append(line)
            index += 1
            continue
        if index + 1 < len(lines) and "|" in stripped and _TABLE_SEPARATOR.match(lines[index + 1]):
            flush_paragraph()
            rows = [_table_cells(stripped)]
            index += 2
            while index < len(lines) and "|" in lines[index] and lines[index].strip():
                rows.append(_table_cells(lines[index]))
                index += 1
            column_count = max(len(row) for row in rows)
            normalized = [row + [""] * (column_count - len(row)) for row in rows]
            rendered = [
                [Paragraph(_inline_markdown(cell), styles["cell_head"] if row_index == 0 else styles["cell"]) for cell in row]
                for row_index, row in enumerate(normalized)
            ]
            usable_width = A4[0] - 36 * mm
            table = Table(rendered, colWidths=[usable_width / column_count] * column_count, repeatRows=1, hAlign="LEFT")
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2f6")),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#cfd6e2")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.extend([table, Spacer(1, 7)])
            continue
        heading = re.match(r"^(#{1,3})\s+(.+)$", stripped)
        if heading:
            flush_paragraph()
            level = len(heading.group(1))
            style = styles["title"] if level == 1 else styles["h2"] if level == 2 else styles["h3"]
            story.append(Paragraph(_inline_markdown(heading.group(2)), style))
        elif re.match(r"^[-*_]{3,}$", stripped):
            flush_paragraph()
            story.append(Spacer(1, 5))
        elif stripped.startswith(">"):
            flush_paragraph()
            story.append(Paragraph(_inline_markdown(stripped.lstrip("> ")), styles["quote"]))
        elif re.match(r"^[-*+]\s+", stripped):
            flush_paragraph()
            story.append(Paragraph(_inline_markdown(re.sub(r"^[-*+]\s+", "", stripped)), styles["bullet"], bulletText="•"))
        elif re.match(r"^\d+[.)]\s+", stripped):
            flush_paragraph()
            number = re.match(r"^(\d+)[.)]", stripped).group(1)
            story.append(Paragraph(_inline_markdown(re.sub(r"^\d+[.)]\s+", "", stripped)), styles["bullet"], bulletText=f"{number}."))
        elif stripped == "":
            flush_paragraph()
        elif stripped == "<!-- pagebreak -->":
            flush_paragraph()
            story.append(PageBreak())
        else:
            paragraph.append(stripped)
        index += 1
    flush_paragraph()
    if code_lines:
        story.append(Preformatted("\n".join(code_lines), styles["code"], maxLineLength=108))
    return story


def _render_pdf(markdown: str, title: str, filename: str) -> bytes:
    font_name, mono_name = _register_fonts()
    styles = _pdf_styles(font_name, mono_name)
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=22 * mm,
        bottomMargin=20 * mm,
        title=title,
        author="THOS Threat Hunting Operations",
        subject=filename,
    )

    def decorate(canvas, doc):
        canvas.saveState()
        canvas.setFont(font_name, 7)
        canvas.setFillColor(colors.HexColor("#667085"))
        canvas.drawString(18 * mm, A4[1] - 12 * mm, "THOS  /  VERIFIED HUNT REPORT")
        canvas.setStrokeColor(colors.HexColor("#dfe4ec"))
        canvas.line(18 * mm, A4[1] - 14 * mm, A4[0] - 18 * mm, A4[1] - 14 * mm)
        canvas.line(18 * mm, 14 * mm, A4[0] - 18 * mm, 14 * mm)
        canvas.setFillColor(colors.HexColor("#98a2b3"))
        canvas.drawString(18 * mm, 9 * mm, filename[:74])
        canvas.drawRightString(A4[0] - 18 * mm, 9 * mm, f"Page {doc.page}")
        canvas.restoreState()

    story = _markdown_story(markdown, styles)
    if not story:
        story = [Paragraph("Empty report", styles["title"])]
    document.build(story, onFirstPage=decorate, onLaterPages=decorate)
    return output.getvalue()


@app.get("/api/reports/{filename}/pdf")
async def download_pdf(filename: str, request: Request):
    control_plane.require_feature(request, "reports")
    path = _safe_report(filename)
    markdown = _clean_report_presentation(
        path.read_text(encoding="utf-8", errors="replace")
    )
    title = _first_heading(markdown, path.stem)
    try:
        content = _render_pdf(markdown, title, path.name)
    except Exception as exc:  # noqa: BLE001 - PDF errors become a clean API response
        logger.exception("report PDF rendering failed")
        raise HTTPException(status_code=500, detail=f"could not render report PDF: {exc}") from exc
    return Response(
        content=content,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{path.stem}.pdf"'},
    )


app.include_router(control_plane.router)


if not STATIC_DIR.is_dir():
    logger.warning("UI static directory is missing: %s", STATIC_DIR)
else:
    assets_dir = STATIC_DIR / "assets"
    if assets_dir.is_dir():
        app.mount("/assets", StaticFiles(directory=assets_dir), name="ui-assets")

    @app.get("/{full_path:path}", include_in_schema=False)
    async def serve_ui_route(full_path: str):
        path = f"/{full_path}".rstrip("/") or "/"
        if not _is_spa_route(path) and path != "/index.html":
            raise HTTPException(status_code=404, detail="page not found")
        return FileResponse(
            STATIC_DIR / "index.html",
            headers={"Cache-Control": "no-store"},
        )
