from io import BytesIO

from openpyxl import load_workbook

from services.api.ui_gateway import (
    LogSearchExportRequest,
    _build_log_search_workbook,
)


def test_log_search_workbook_contains_logs_metadata_and_mapping():
    payload = LogSearchExportRequest(
        portable_query='process_name == "powershell.exe"',
        query='{"query":{"match_phrase":{"process.name":"powershell.exe"}}}',
        siem_type="elasticsearch",
        lookback_minutes=60,
        executed_at="2026-08-02T12:00:00+05:30",
        field_mapping={"process_name": "process.name"},
        logs=[{
            "@timestamp": "2026-08-02T06:30:00Z",
            "host": {"name": "endpoint-01"},
            "process": {"name": "powershell.exe"},
            "untrusted": "=HYPERLINK(\"https://example.invalid\")",
        }],
    )

    workbook = load_workbook(BytesIO(_build_log_search_workbook(payload)))

    assert workbook.sheetnames == ["Logs", "Search metadata", "Field mapping"]
    logs = workbook["Logs"]
    headers = [cell.value for cell in logs[1]]
    assert "@timestamp" in headers
    assert "host.name" in headers
    assert "process.name" in headers
    unsafe_column = headers.index("untrusted") + 1
    assert logs.cell(2, unsafe_column).value.startswith("'=")
    assert logs.freeze_panes == "A2"
    assert workbook["Search metadata"]["B2"].value == "elasticsearch"
    assert workbook["Field mapping"]["A2"].value == "process_name"
    assert workbook["Field mapping"]["B2"].value == "process.name"


def test_empty_log_search_export_remains_a_valid_workbook():
    payload = LogSearchExportRequest(
        portable_query="failed authentication",
        query="failed, authentication",
        siem_type="folder",
        logs=[],
    )

    workbook = load_workbook(BytesIO(_build_log_search_workbook(payload)))

    assert workbook["Logs"]["A1"].value == "No records returned"
    assert workbook["Search metadata"]["B8"].value == 0
